import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

with open("store-report-preview.json", encoding="utf-8") as f:
    report = json.load(f)

daily = report["daily"]
months = report["months"]

HEADERS = ["월", "주차", "방문자 수", "영수증 건수", "전환율", "총 판매 수량", "총 정상가 합산",
           "총 판매금액", "매출 할인", "총 실결제금액", "목표 매출", "매출 달성율",
           "외국인 매출", "외국인 비중", "AOV", "UPT"]
N_COLS = len(HEADERS)

CURRENCY_COLS = {6, 7, 8, 9, 10, 11, 13, 15}  # 1-indexed columns with amounts
PERCENT_COLS = {5, 12, 14}
DECIMAL_COLS = {16}

FONT_NAME = "Arial"
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="000000")
TOTAL_FONT = Font(name=FONT_NAME, bold=True)
TOTAL_FILL = PatternFill("solid", start_color="D9D9D9")
MONTH_FONT = Font(name=FONT_NAME, bold=True)
MONTH_FILL = PatternFill("solid", start_color="F2F2F2")
NORMAL_FONT = Font(name=FONT_NAME)
NOTE_FONT = Font(name=FONT_NAME, italic=True, color="FF0000", size=9)

wb = Workbook()
ws = wb.active
ws.title = "애월"

# 요약 섹션과 raw data 섹션 각각의 시작 행을 먼저 계산 (raw data가 뒤에 오지만 요약 수식이 그 범위를 참조)
n_month_rows = sum(1 + len(m["weeks"]) for m in months)
summary_header_row = 1
total_row = 2
summary_start = 3
summary_end = summary_start + n_month_rows - 1
raw_header_row = summary_end + 2
raw_title_row = raw_header_row + 1
raw_header2_row = raw_title_row + 1
raw_start = raw_header2_row + 1
raw_end = raw_start + len(daily) - 1


def write_header(row):
    for c, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=row, column=c, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def apply_number_format(cell, col):
    if col in CURRENCY_COLS:
        cell.number_format = "#,##0;(#,##0);-"
    elif col in PERCENT_COLS:
        cell.number_format = "0.00%"
    elif col in DECIMAL_COLS:
        cell.number_format = "0.0"


def write_summary_row(row, month_label, week_label, r1, r2, font, fill):
    ws.cell(row=row, column=1, value=month_label)
    ws.cell(row=row, column=2, value=week_label)
    visitors_col, receipts_col = 3, 4
    qty_col, tag_col, sales_col, disc_col, pay_col = 6, 7, 8, 9, 10
    foreign_col = 13
    col_letters = {c: get_column_letter(c) for c in range(1, N_COLS + 1)}

    ws.cell(row=row, column=visitors_col, value=f"=SUM({col_letters[visitors_col]}{r1}:{col_letters[visitors_col]}{r2})")
    ws.cell(row=row, column=receipts_col, value=f"=SUM({col_letters[receipts_col]}{r1}:{col_letters[receipts_col]}{r2})")
    ws.cell(row=row, column=5, value=f"=IFERROR({col_letters[receipts_col]}{row}/{col_letters[visitors_col]}{row},0)")
    ws.cell(row=row, column=qty_col, value=f"=SUM({col_letters[qty_col]}{r1}:{col_letters[qty_col]}{r2})")
    ws.cell(row=row, column=tag_col, value=f"=SUM({col_letters[tag_col]}{r1}:{col_letters[tag_col]}{r2})")
    ws.cell(row=row, column=sales_col, value=f"=SUM({col_letters[sales_col]}{r1}:{col_letters[sales_col]}{r2})")
    ws.cell(row=row, column=disc_col, value=f"=SUM({col_letters[disc_col]}{r1}:{col_letters[disc_col]}{r2})")
    ws.cell(row=row, column=pay_col, value=f"=SUM({col_letters[pay_col]}{r1}:{col_letters[pay_col]}{r2})")
    ws.cell(row=row, column=11, value=None)  # 목표 매출: 수기 입력
    ws.cell(row=row, column=12, value=f"=IFERROR({col_letters[pay_col]}{row}/K{row},0)")
    ws.cell(row=row, column=foreign_col, value=f"=SUM({col_letters[foreign_col]}{r1}:{col_letters[foreign_col]}{r2})")
    ws.cell(row=row, column=14, value=f"=IFERROR({col_letters[foreign_col]}{row}/{col_letters[pay_col]}{row},0)")
    ws.cell(row=row, column=15, value=f"=IFERROR({col_letters[pay_col]}{row}/{col_letters[receipts_col]}{row},0)")
    ws.cell(row=row, column=16, value=f"=IFERROR({col_letters[qty_col]}{row}/{col_letters[receipts_col]}{row},0)")

    for c in range(1, N_COLS + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        if fill is not None:
            cell.fill = fill
        apply_number_format(cell, c)


write_header(summary_header_row)
write_summary_row(total_row, f"{daily[0]['date'][:4]} TTL", "", raw_start, raw_end, TOTAL_FONT, TOTAL_FILL)

row = summary_start
for m in months:
    month_num = int(m["month"].split("-")[1])
    week_rows_in_month = [d for d in daily if d["date"][:7] == m["month"]]
    first_date_in_month = week_rows_in_month[0]["date"]
    last_date_in_month = week_rows_in_month[-1]["date"]
    r1 = raw_start + [d["date"] for d in daily].index(first_date_in_month)
    r2 = raw_start + [d["date"] for d in daily].index(last_date_in_month)
    write_summary_row(row, f"{month_num}월 합산", "", r1, r2, MONTH_FONT, MONTH_FILL)
    row += 1
    for w in m["weeks"]:
        week_dates = [d["date"] for d in daily if d["date"][:7] == m["month"] and d["week"] == w["week"]]
        wr1 = raw_start + [d["date"] for d in daily].index(week_dates[0])
        wr2 = raw_start + [d["date"] for d in daily].index(week_dates[-1])
        write_summary_row(row, "", f"{w['week']}주차", wr1, wr2, NORMAL_FONT, None)
        row += 1

write_header(raw_header_row)
title_cell = ws.cell(row=raw_title_row, column=1, value=f"{daily[0]['date'][:4]} Raw Data")
title_cell.font = MONTH_FONT
write_header(raw_header2_row)

for i, d in enumerate(daily):
    r = raw_start + i
    ws.cell(row=r, column=1, value=d["date"])
    ws.cell(row=r, column=2, value=f"{d['week']}주차")
    ws.cell(row=r, column=3, value=d["visitors"])
    ws.cell(row=r, column=4, value=d["receiptCount"])
    ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/C{r},0)")
    ws.cell(row=r, column=6, value=d["totalQty"])
    ws.cell(row=r, column=7, value=d["totalTagPrice"])
    ws.cell(row=r, column=8, value=d["totalSalesPrice"])
    ws.cell(row=r, column=9, value=d["discountAmount"])
    ws.cell(row=r, column=10, value=d["totalPaymentAmount"])
    ws.cell(row=r, column=11, value=None)
    ws.cell(row=r, column=12, value=f"=IFERROR(J{r}/K{r},0)")
    ws.cell(row=r, column=13, value=d["foreignSalesAmount"])
    ws.cell(row=r, column=14, value=f"=IFERROR(M{r}/J{r},0)")
    ws.cell(row=r, column=15, value=f"=IFERROR(J{r}/D{r},0)")
    ws.cell(row=r, column=16, value=f"=IFERROR(F{r}/D{r},0)")
    for c in range(1, N_COLS + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = NORMAL_FONT
        apply_number_format(cell, c)

note_row = raw_end + 2
note_cell = ws.cell(row=note_row, column=1,
                     value="※ 목표 매출/매출 달성율은 API로 산출할 수 없는 값이라 비워뒀습니다 (수기 입력 필요).")
note_cell.font = NOTE_FONT

for c in range(1, N_COLS + 1):
    ws.column_dimensions[get_column_letter(c)].width = 14
ws.column_dimensions["A"].width = 12

wb.save("reports/Verish_애월_pilot.xlsx")
print("saved. summary rows:", summary_start, "-", summary_end, "raw rows:", raw_start, "-", raw_end)
