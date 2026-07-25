import csv
import sys
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

CSV_PATH = sys.argv[1]
OUT_PATH = sys.argv[2]

FONT_NAME = "Arial"
BLUE = Font(name=FONT_NAME, color="0000FF")
BLACK = Font(name=FONT_NAME, color="000000")
BOLD = Font(name=FONT_NAME, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
YELLOW = PatternFill("solid", fgColor="FFFF00")
DATE_FMT = "yyyy-mm-dd"
WON_FMT = "#,##0"
PCT_FMT = "0.0%"


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# ---------------- 가정 (Assumptions) ----------------
ws_a = wb.active
ws_a.title = "가정"
ws_a.append(["항목", "값", "비고"])
style_header(ws_a, 1, 3)

rows_a = [
    ("8월 목표매출 (자사몰/CAFE24)", 4177348798, "사용자 제공 값"),
    ("요일비중 산출기간 시작", date(2026, 6, 1), ""),
    ("요일비중 산출기간 종료", date(2026, 7, 13), "7/14 이후는 카드 정산 지연으로 매출이 0에 수렴, 요일비중 산출에서 제외"),
    ("1차 프로모션 시작", date(2026, 8, 3), ""),
    ("1차 프로모션 종료", date(2026, 8, 14), ""),
    ("1차 프로모션 업리프트 배수", 1.5, "\"매출 150%로 업리프트\" = 평시 대비 1.5배로 해석"),
    ("2차 프로모션 시작", date(2026, 8, 18), "사용자 확인: 원 요청 \"8.18~8.14\"는 날짜 역순이라 8.18~8.24로 확인받음"),
    ("2차 프로모션 종료", date(2026, 8, 24), ""),
    ("2차 프로모션 업리프트 배수", 4.0, "\"매출 400%로 업리프트\" = 평시 대비 4배로 해석"),
]
for r in rows_a:
    ws_a.append(list(r))

for i, r in enumerate(range(2, 2 + len(rows_a)), start=0):
    ws_a.cell(row=r, column=1).font = BLACK
    val_cell = ws_a.cell(row=r, column=2)
    val_cell.font = BLUE
    val_cell.fill = YELLOW
    note_cell = ws_a.cell(row=r, column=3)
    note_cell.font = BLACK
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    if isinstance(rows_a[i][1], date):
        val_cell.number_format = DATE_FMT
    elif rows_a[i][0].endswith("매출"):
        val_cell.number_format = WON_FMT
    elif "배수" in rows_a[i][0]:
        val_cell.number_format = "0.0\"x\""

autofit(ws_a, [30, 16, 70])
ws_a.freeze_panes = "A2"

# ---------------- 원천데이터 (raw CAFE24 daily sales) ----------------
ws_r = wb.create_sheet("원천데이터(CAFE24)")
ws_r.append(["날짜", "요일번호", "요일", "실결제금액(원)", "요일비중 산출 포함여부"])
style_header(ws_r, 1, 5)

with open(CSV_PATH, encoding="utf-8") as f:
    reader = csv.DictReader(f)
    src_rows = list(reader)

r0 = 2
for i, rec in enumerate(src_rows):
    d = date(int(rec["date"][0:4]), int(rec["date"][4:6]), int(rec["date"][6:8]))
    row = r0 + i
    ws_r.cell(row=row, column=1, value=d).number_format = DATE_FMT
    ws_r.cell(row=row, column=1).font = BLUE
    ws_r.cell(row=row, column=2, value=f"=WEEKDAY(A{row},1)-1").font = BLACK
    ws_r.cell(row=row, column=3, value=f'=CHOOSE(B{row}+1,"일","월","화","수","목","금","토")').font = BLACK
    amt_cell = ws_r.cell(row=row, column=4, value=int(rec["amount"]))
    amt_cell.font = BLUE
    amt_cell.number_format = WON_FMT
    ws_r.cell(
        row=row,
        column=5,
        value=f'=IF(AND(A{row}>=가정!$B$3,A{row}<=가정!$B$4),"포함","제외(정산지연)")',
    ).font = BLACK

last_src_row = r0 + len(src_rows) - 1
autofit(ws_r, [14, 10, 8, 18, 22])
ws_r.freeze_panes = "A2"

# ---------------- 요일별비중 (weekday mix) ----------------
ws_w = wb.create_sheet("요일별비중")
ws_w.append(["요일번호", "요일", "평균매출(원)", "비중(%)"])
style_header(ws_w, 1, 4)

for dow in range(7):
    row = 2 + dow
    ws_w.cell(row=row, column=1, value=dow).font = BLUE
    ws_w.cell(row=row, column=2, value=f'=CHOOSE(A{row}+1,"일","월","화","수","목","금","토")').font = BLACK
    avg_formula = (
        f"=AVERAGEIFS('원천데이터(CAFE24)'!$D$2:$D${last_src_row},"
        f"'원천데이터(CAFE24)'!$B$2:$B${last_src_row},A{row},"
        f"'원천데이터(CAFE24)'!$E$2:$E${last_src_row},\"포함\")"
    )
    c = ws_w.cell(row=row, column=3, value=avg_formula)
    c.font = BLACK
    c.number_format = WON_FMT

total_row = 9
ws_w.cell(row=total_row, column=2, value="합계").font = BOLD
sum_cell = ws_w.cell(row=total_row, column=3, value=f"=SUM(C2:C8)")
sum_cell.font = BOLD
sum_cell.number_format = WON_FMT

for dow in range(7):
    row = 2 + dow
    pct_cell = ws_w.cell(row=row, column=4, value=f"=C{row}/$C${total_row}")
    pct_cell.font = BLACK
    pct_cell.number_format = PCT_FMT

pct_total = ws_w.cell(row=total_row, column=4, value=f"=SUM(D2:D8)")
pct_total.font = BOLD
pct_total.number_format = PCT_FMT

autofit(ws_w, [10, 8, 18, 10])
ws_w.freeze_panes = "A2"

# ---------------- 일별배분 (August daily allocation) ----------------
ws_d = wb.create_sheet("일별배분")
ws_d.append(["날짜", "요일번호", "요일", "구간", "업리프트 배수", "기준지수(요일 평균매출)", "조정지수", "배분금액(원)"])
style_header(ws_d, 1, 8)

start = date(2026, 8, 1)
n_days = 31
first_row = 2
last_row = first_row + n_days - 1

for i in range(n_days):
    d = start + timedelta(days=i)
    row = first_row + i
    ws_d.cell(row=row, column=1, value=d).number_format = DATE_FMT
    ws_d.cell(row=row, column=1).font = BLUE
    ws_d.cell(row=row, column=2, value=f"=WEEKDAY(A{row},1)-1").font = BLACK
    ws_d.cell(row=row, column=3, value=f'=CHOOSE(B{row}+1,"일","월","화","수","목","금","토")').font = BLACK

    segment_formula = (
        f'=IF(AND(A{row}>=가정!$B$5,A{row}<=가정!$B$6),"1차 프로모션",'
        f'IF(AND(A{row}>=가정!$B$8,A{row}<=가정!$B$9),"2차 프로모션","평시"))'
    )
    ws_d.cell(row=row, column=4, value=segment_formula).font = BLACK

    mult_formula = (
        f'=IF(AND(A{row}>=가정!$B$5,A{row}<=가정!$B$6),가정!$B$7,'
        f'IF(AND(A{row}>=가정!$B$8,A{row}<=가정!$B$9),가정!$B$10,1))'
    )
    ws_d.cell(row=row, column=5, value=mult_formula).font = BLACK
    ws_d.cell(row=row, column=5).number_format = "0.0\"x\""

    base_idx_formula = f"=INDEX('요일별비중'!$C$2:$C$8,MATCH(B{row},'요일별비중'!$A$2:$A$8,0))"
    c6 = ws_d.cell(row=row, column=6, value=base_idx_formula)
    c6.font = BLACK
    c6.number_format = WON_FMT

    adj_idx_formula = f"=F{row}*E{row}"
    c7 = ws_d.cell(row=row, column=7, value=adj_idx_formula)
    c7.font = BLACK
    c7.number_format = WON_FMT

    if i < n_days - 1:
        amt_formula = f"=ROUND(G{row}/SUM($G${first_row}:$G${last_row})*가정!$B$2,0)"
    else:
        # last day absorbs the rounding remainder so the total matches the fixed target exactly
        amt_formula = f"=가정!$B$2-SUM(H{first_row}:H{last_row-1})"
    c8 = ws_d.cell(row=row, column=8, value=amt_formula)
    c8.font = BLACK
    c8.number_format = WON_FMT

sum_row = last_row + 1
ws_d.cell(row=sum_row, column=3, value="합계").font = BOLD
tot_cell = ws_d.cell(row=sum_row, column=8, value=f"=SUM(H{first_row}:H{last_row})")
tot_cell.font = BOLD
tot_cell.number_format = WON_FMT

check_row = sum_row + 1
ws_d.cell(row=check_row, column=3, value="목표 대비 차액(검증, 0이어야 함)").font = BLACK
diff_cell = ws_d.cell(row=check_row, column=8, value=f"=H{sum_row}-가정!$B$2")
diff_cell.font = BLACK
diff_cell.number_format = WON_FMT

promo1_row = check_row + 2
ws_d.cell(row=promo1_row, column=3, value="1차 프로모션 합계").font = BLACK
p1 = ws_d.cell(row=promo1_row, column=8, value=f'=SUMIF($D${first_row}:$D${last_row},"1차 프로모션",$H${first_row}:$H${last_row})')
p1.font = BLACK
p1.number_format = WON_FMT

promo2_row = promo1_row + 1
ws_d.cell(row=promo2_row, column=3, value="2차 프로모션 합계").font = BLACK
p2 = ws_d.cell(row=promo2_row, column=8, value=f'=SUMIF($D${first_row}:$D${last_row},"2차 프로모션",$H${first_row}:$H${last_row})')
p2.font = BLACK
p2.number_format = WON_FMT

rest_row = promo2_row + 1
ws_d.cell(row=rest_row, column=3, value="평시 합계").font = BLACK
p3 = ws_d.cell(row=rest_row, column=8, value=f'=SUMIF($D${first_row}:$D${last_row},"평시",$H${first_row}:$H${last_row})')
p3.font = BLACK
p3.number_format = WON_FMT

autofit(ws_d, [14, 10, 8, 14, 12, 20, 16, 18])
ws_d.freeze_panes = "A2"

# reorder sheets: 가정, 일별배분, 요일별비중, 원천데이터
wb._sheets = [ws_a, ws_d, ws_w, ws_r]

# no LibreOffice available on this machine to pre-compute cached formula
# values, so force Excel to fully recalculate the instant the file opens.
wb.calculation.fullCalcOnLoad = True

wb.save(OUT_PATH)
print("saved", OUT_PATH)
