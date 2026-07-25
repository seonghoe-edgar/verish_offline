import csv
import sys
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

TOTAL_CSV = sys.argv[1]
US_CSV = sys.argv[2]
OUT_PATH = sys.argv[3]

FONT_NAME = "Arial"
BLUE = Font(name=FONT_NAME, color="0000FF")
BLACK = Font(name=FONT_NAME, color="000000")
BOLD = Font(name=FONT_NAME, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
YELLOW = PatternFill("solid", fgColor="FFFF00")
DATE_FMT = "yyyy-mm-dd"
WON_FMT = "#,##0"
USD_FMT = "#,##0.00"
PCT_FMT = "0.0%"


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def load_csv(path):
    with open(path, encoding="utf-8") as f:
        return {rec["date"]: float(rec["amount"]) for rec in csv.DictReader(f)}


total_map = load_csv(TOTAL_CSV)
us_map = load_csv(US_CSV)
dates = sorted(total_map.keys())

wb = Workbook()

# ---------------- 가정 ----------------
ws_a = wb.active
ws_a.title = "가정"
ws_a.append(["항목", "값", "비고"])
style_header(ws_a, 1, 3)

rows_a = [
    ("8월 목표매출 (해외자사몰, 원화)", 3125608605, "사용자 제공 값"),
    ("요일비중 산출기간 시작", date(2026, 6, 1), "Shopify(verish-int.myshopify.com) 커넥터, ShopifyQL total_sales 기준"),
    ("요일비중 산출기간 종료", date(2026, 7, 24), "정산 지연 없음 — 전체 기간 사용"),
    ("프로모션 시작", date(2026, 8, 10), ""),
    ("프로모션 종료", date(2026, 8, 23), ""),
    (
        "프로모션 업리프트 배수 (미국외 전용)",
        1.8,
        "\"매출 180%로 업리프트\" = 평시 대비 1.8배. 사용자 확인: 미국은 프로모션 영향이 없어 미국 매출에는 배수를 적용하지 않고 평시 요일 지수를 그대로 사용",
    ),
]
for r in rows_a:
    ws_a.append(list(r))

for i, r in enumerate(range(2, 2 + len(rows_a)), start=0):
    ws_a.cell(row=r, column=1).font = BLACK
    val_cell = ws_a.cell(row=r, column=2)
    val_cell.font = BLUE
    val_cell.fill = YELLOW
    note_cell = ws_a.cell(row=r, column=3)
    note_cell.font = BLACK
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    if isinstance(rows_a[i][1], date):
        val_cell.number_format = DATE_FMT
    elif "원화" in rows_a[i][0]:
        val_cell.number_format = WON_FMT
    elif "배수" in rows_a[i][0]:
        val_cell.number_format = "0.0\"x\""

autofit(ws_a, [34, 16, 75])
ws_a.freeze_panes = "A2"

# ---------------- 원천데이터 (Shopify daily USD sales, 미국 vs 미국외) ----------------
ws_r = wb.create_sheet("원천데이터(Shopify)")
ws_r.append(["날짜", "요일번호", "요일", "전체(USD)", "미국(USD)", "미국외(USD)"])
style_header(ws_r, 1, 6)

r0 = 2
for i, dstr in enumerate(dates):
    y, m, d_ = dstr.split("-")
    d = date(int(y), int(m), int(d_))
    row = r0 + i
    ws_r.cell(row=row, column=1, value=d).number_format = DATE_FMT
    ws_r.cell(row=row, column=1).font = BLUE
    ws_r.cell(row=row, column=2, value=f"=WEEKDAY(A{row},1)-1").font = BLACK
    ws_r.cell(row=row, column=3, value=f'=CHOOSE(B{row}+1,"일","월","화","수","목","금","토")').font = BLACK

    total_cell = ws_r.cell(row=row, column=4, value=total_map[dstr])
    total_cell.font = BLUE
    total_cell.number_format = USD_FMT

    us_cell = ws_r.cell(row=row, column=5, value=us_map[dstr])
    us_cell.font = BLUE
    us_cell.number_format = USD_FMT

    non_us_cell = ws_r.cell(row=row, column=6, value=f"=D{row}-E{row}")
    non_us_cell.font = BLACK
    non_us_cell.number_format = USD_FMT

last_src_row = r0 + len(dates) - 1
autofit(ws_r, [14, 10, 8, 14, 14, 14])
ws_r.freeze_panes = "A2"

# ---------------- 요일별비중 (separate US / non-US weekday baselines) ----------------
ws_w = wb.create_sheet("요일별비중")
ws_w.append(["요일번호", "요일", "미국 평균매출(USD)", "미국외 평균매출(USD)"])
style_header(ws_w, 1, 4)

for dow in range(7):
    row = 2 + dow
    ws_w.cell(row=row, column=1, value=dow).font = BLUE
    ws_w.cell(row=row, column=2, value=f'=CHOOSE(A{row}+1,"일","월","화","수","목","금","토")').font = BLACK
    us_avg_formula = (
        f"=AVERAGEIF('원천데이터(Shopify)'!$B$2:$B${last_src_row},A{row},"
        f"'원천데이터(Shopify)'!$E$2:$E${last_src_row})"
    )
    c = ws_w.cell(row=row, column=3, value=us_avg_formula)
    c.font = BLACK
    c.number_format = USD_FMT

    non_us_avg_formula = (
        f"=AVERAGEIF('원천데이터(Shopify)'!$B$2:$B${last_src_row},A{row},"
        f"'원천데이터(Shopify)'!$F$2:$F${last_src_row})"
    )
    d_cell = ws_w.cell(row=row, column=4, value=non_us_avg_formula)
    d_cell.font = BLACK
    d_cell.number_format = USD_FMT

autofit(ws_w, [10, 8, 20, 20])
ws_w.freeze_panes = "A2"

# ---------------- 일별배분 ----------------
ws_d = wb.create_sheet("일별배분")
ws_d.append(
    [
        "날짜", "요일번호", "요일", "구간",
        "업리프트 배수(미국외 전용)",
        "미국 기준지수(USD, 배수 미적용)",
        "미국외 기준지수(USD)",
        "미국외 조정지수",
        "합계지수",
        "미국 배분금액(원)",
        "미국외 배분금액(원)",
        "합계 배분금액(원)",
    ]
)
style_header(ws_d, 1, 12)

start = date(2026, 8, 1)
n_days = 31
first_row = 2
last_row = first_row + n_days - 1

for i in range(n_days):
    d = start + timedelta(days=i)
    row = first_row + i
    ws_d.cell(row=row, column=1, value=d).number_format = DATE_FMT
    ws_d.cell(row=row, column=1).font = BLUE
    ws_d.cell(row=row, column=2, value=f"=WEEKDAY(A{row},1)-1").font = BLACK
    ws_d.cell(row=row, column=3, value=f'=CHOOSE(B{row}+1,"일","월","화","수","목","금","토")').font = BLACK

    segment_formula = f'=IF(AND(A{row}>=가정!$B$5,A{row}<=가정!$B$6),"프로모션","평시")'
    ws_d.cell(row=row, column=4, value=segment_formula).font = BLACK

    mult_formula = f'=IF(AND(A{row}>=가정!$B$5,A{row}<=가정!$B$6),가정!$B$7,1)'
    ws_d.cell(row=row, column=5, value=mult_formula).font = BLACK
    ws_d.cell(row=row, column=5).number_format = "0.0\"x\""

    us_idx_formula = f"=INDEX('요일별비중'!$C$2:$C$8,MATCH(B{row},'요일별비중'!$A$2:$A$8,0))"
    c6 = ws_d.cell(row=row, column=6, value=us_idx_formula)
    c6.font = BLACK
    c6.number_format = USD_FMT

    non_us_idx_formula = f"=INDEX('요일별비중'!$D$2:$D$8,MATCH(B{row},'요일별비중'!$A$2:$A$8,0))"
    c7 = ws_d.cell(row=row, column=7, value=non_us_idx_formula)
    c7.font = BLACK
    c7.number_format = USD_FMT

    non_us_adj_formula = f"=G{row}*E{row}"
    c8 = ws_d.cell(row=row, column=8, value=non_us_adj_formula)
    c8.font = BLACK
    c8.number_format = USD_FMT

    total_idx_formula = f"=F{row}+H{row}"
    c9 = ws_d.cell(row=row, column=9, value=total_idx_formula)
    c9.font = BLACK
    c9.number_format = USD_FMT

    # US never gets the uplift, but it's still rescaled (along with non-US)
    # against the *combined* index total so the fixed monthly target holds.
    us_amt_formula = f"=ROUND(F{row}/SUM($I${first_row}:$I${last_row})*가정!$B$2,0)"
    c10 = ws_d.cell(row=row, column=10, value=us_amt_formula)
    c10.font = BLACK
    c10.number_format = WON_FMT

    if i < n_days - 1:
        non_us_amt_formula = f"=ROUND(H{row}/SUM($I${first_row}:$I${last_row})*가정!$B$2,0)"
    else:
        # last day's non-US amount absorbs the rounding remainder so the
        # grand total (US + non-US, all 31 days) matches the fixed target exactly
        non_us_amt_formula = (
            f"=가정!$B$2-SUM(J{first_row}:J{last_row})-SUM(K{first_row}:K{last_row-1})"
        )
    c11 = ws_d.cell(row=row, column=11, value=non_us_amt_formula)
    c11.font = BLACK
    c11.number_format = WON_FMT

    c12 = ws_d.cell(row=row, column=12, value=f"=J{row}+K{row}")
    c12.font = BLACK
    c12.number_format = WON_FMT

sum_row = last_row + 1
ws_d.cell(row=sum_row, column=3, value="합계").font = BOLD
for col in (10, 11, 12):
    tot_cell = ws_d.cell(row=sum_row, column=col, value=f"=SUM({get_column_letter(col)}{first_row}:{get_column_letter(col)}{last_row})")
    tot_cell.font = BOLD
    tot_cell.number_format = WON_FMT

check_row = sum_row + 1
ws_d.cell(row=check_row, column=3, value="목표 대비 차액(검증, 0이어야 함)").font = BLACK
diff_cell = ws_d.cell(row=check_row, column=12, value=f"=L{sum_row}-가정!$B$2")
diff_cell.font = BLACK
diff_cell.number_format = WON_FMT

promo_row = check_row + 2
ws_d.cell(row=promo_row, column=3, value="프로모션 구간 — 미국 합계 (업리프트 미적용)").font = BLACK
p1 = ws_d.cell(row=promo_row, column=10, value=f'=SUMIF($D${first_row}:$D${last_row},"프로모션",$J${first_row}:$J${last_row})')
p1.font = BLACK
p1.number_format = WON_FMT

promo_row2 = promo_row + 1
ws_d.cell(row=promo_row2, column=3, value="프로모션 구간 — 미국외 합계 (업리프트 적용)").font = BLACK
p2 = ws_d.cell(row=promo_row2, column=11, value=f'=SUMIF($D${first_row}:$D${last_row},"프로모션",$K${first_row}:$K${last_row})')
p2.font = BLACK
p2.number_format = WON_FMT

rest_row = promo_row2 + 1
ws_d.cell(row=rest_row, column=3, value="평시 구간 합계").font = BLACK
p3 = ws_d.cell(row=rest_row, column=12, value=f'=SUMIF($D${first_row}:$D${last_row},"평시",$L${first_row}:$L${last_row})')
p3.font = BLACK
p3.number_format = WON_FMT

autofit(ws_d, [14, 10, 8, 12, 18, 22, 20, 16, 16, 16, 18, 18])
ws_d.freeze_panes = "A2"

wb._sheets = [ws_a, ws_d, ws_w, ws_r]
wb.calculation.fullCalcOnLoad = True
wb.save(OUT_PATH)
print("saved", OUT_PATH)
