import csv
import sys
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

CSV_PATH = sys.argv[1]  # cafe24_daily.csv (reused for weekday mix)
OUT_PATH = sys.argv[2]

FONT_NAME = "Arial"
BLUE = Font(name=FONT_NAME, color="0000FF")
BLACK = Font(name=FONT_NAME, color="000000")
BOLD = Font(name=FONT_NAME, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
YELLOW = PatternFill("solid", fgColor="FFFF00")
DATE_FMT = "yyyy-mm-dd"
WON_FMT = "#,##0"
PCT_FMT = "0.0%"


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# ---------------- 채널목록 (reference only, not used in the calc) ----------------
ws_c = wb.active
ws_c.title = "채널목록(참고)"
ws_c.append(["채널", "Final Round 매출(원)", "비고"])
style_header(ws_c, 1, 3)
channels = [
    ("지그재그", 329000000, "8/3 라이브 방송 별도 반영"),
    ("29CM", 409000000, ""),
    ("무신사", 236000000, ""),
    ("에이블리", 114000000, ""),
    ("스마트스토어", 113000000, ""),
    ("카카오 선물하기", 31000000, ""),
    ("Wconcept", 0, ""),
    ("올리브영_온라인", 19000000, ""),
    ("올리브영_오프라인", 187000000, ""),
    ("면세점_롯데", 159000000, ""),
    ("면세점_신라", 99000000, ""),
    ("면세점_신세계", 127000000, ""),
]
for i, (name, amt, note) in enumerate(channels):
    row = 2 + i
    ws_c.cell(row=row, column=1, value=name).font = BLUE
    c2 = ws_c.cell(row=row, column=2, value=amt)
    c2.font = BLUE
    c2.number_format = WON_FMT
    ws_c.cell(row=row, column=3, value=note).font = BLACK
sum_row = 2 + len(channels)
ws_c.cell(row=sum_row, column=1, value="합계 (자사몰 제외)").font = BOLD
tot = ws_c.cell(row=sum_row, column=2, value=f"=SUM(B2:B{sum_row - 1})")
tot.font = BOLD
tot.number_format = WON_FMT
autofit(ws_c, [20, 20, 30])
ws_c.freeze_panes = "A2"

# ---------------- 가정 ----------------
ws_a = wb.create_sheet("가정")
ws_a.append(["항목", "값", "비고"])
style_header(ws_a, 1, 3)
rows_a = [
    ("8월 목표매출 (국내 외부몰 합산, 자사몰 제외)", f"='채널목록(참고)'!B{sum_row}", "12개 채널 Final Round 합계"),
    ("요일비중 산출 기준", "자사몰(CAFE24) 요일별비중 재사용", "6/1~7/13, 정산지연으로 7/14 이후 제외 (기존 자사몰 분석과 동일)"),
    ("고정 반영 일자", date(2026, 8, 3), "지그재그 단독 라이브 방송일"),
    ("고정 반영 금액(원)", 47000000, "사용자 제공 값 — 이 날은 요일비중 배분 대신 고정액 사용"),
]
for r in rows_a:
    ws_a.append(list(r))
for i, r in enumerate(range(2, 2 + len(rows_a)), start=0):
    ws_a.cell(row=r, column=1).font = BLACK
    val_cell = ws_a.cell(row=r, column=2)
    note_cell = ws_a.cell(row=r, column=3)
    note_cell.font = BLACK
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    if isinstance(rows_a[i][1], date):
        val_cell.font = BLUE
        val_cell.fill = YELLOW
        val_cell.number_format = DATE_FMT
    elif isinstance(rows_a[i][1], (int, float)):
        val_cell.font = BLUE
        val_cell.fill = YELLOW
        val_cell.number_format = WON_FMT
    elif str(rows_a[i][1]).startswith("="):
        val_cell.font = BLACK
        val_cell.number_format = WON_FMT
    else:
        val_cell.font = BLACK
autofit(ws_a, [40, 30, 65])
ws_a.freeze_panes = "A2"

# ---------------- 원천데이터(자사몰, 요일비중 재사용) ----------------
ws_r = wb.create_sheet("원천데이터(자사몰)")
ws_r.append(["날짜", "요일번호", "요일", "실결제금액(원)", "요일비중 산출 포함여부"])
style_header(ws_r, 1, 5)

with open(CSV_PATH, encoding="utf-8") as f:
    src_rows = list(csv.DictReader(f))

CLEAN_FROM = "20260601"
CLEAN_TO = "20260713"
r0 = 2
for i, rec in enumerate(src_rows):
    d = date(int(rec["date"][0:4]), int(rec["date"][4:6]), int(rec["date"][6:8]))
    row = r0 + i
    ws_r.cell(row=row, column=1, value=d).number_format = DATE_FMT
    ws_r.cell(row=row, column=1).font = BLUE
    ws_r.cell(row=row, column=2, value=f"=WEEKDAY(A{row},1)-1").font = BLACK
    ws_r.cell(row=row, column=3, value=f'=CHOOSE(B{row}+1,"일","월","화","수","목","금","토")').font = BLACK
    amt_cell = ws_r.cell(row=row, column=4, value=int(rec["amount"]))
    amt_cell.font = BLUE
    amt_cell.number_format = WON_FMT
    included = CLEAN_FROM <= rec["date"] <= CLEAN_TO
    ws_r.cell(row=row, column=5, value="포함" if included else "제외(정산지연)").font = BLACK

last_src_row = r0 + len(src_rows) - 1
autofit(ws_r, [14, 10, 8, 18, 22])
ws_r.freeze_panes = "A2"

# ---------------- 요일별비중 ----------------
ws_w = wb.create_sheet("요일별비중")
ws_w.append(["요일번호", "요일", "평균매출(원, 자사몰기준)", "비중(%)"])
style_header(ws_w, 1, 4)
for dow in range(7):
    row = 2 + dow
    ws_w.cell(row=row, column=1, value=dow).font = BLUE
    ws_w.cell(row=row, column=2, value=f'=CHOOSE(A{row}+1,"일","월","화","수","목","금","토")').font = BLACK
    avg_formula = (
        f"=AVERAGEIFS('원천데이터(자사몰)'!$D$2:$D${last_src_row},"
        f"'원천데이터(자사몰)'!$B$2:$B${last_src_row},A{row},"
        f"'원천데이터(자사몰)'!$E$2:$E${last_src_row},\"포함\")"
    )
    c = ws_w.cell(row=row, column=3, value=avg_formula)
    c.font = BLACK
    c.number_format = WON_FMT
w_total_row = 9
ws_w.cell(row=w_total_row, column=2, value="합계").font = BOLD
w_tot = ws_w.cell(row=w_total_row, column=3, value="=SUM(C2:C8)")
w_tot.font = BOLD
w_tot.number_format = WON_FMT
for dow in range(7):
    row = 2 + dow
    pct = ws_w.cell(row=row, column=4, value=f"=C{row}/$C${w_total_row}")
    pct.font = BLACK
    pct.number_format = PCT_FMT
autofit(ws_w, [10, 8, 22, 10])
ws_w.freeze_panes = "A2"

# ---------------- 일별배분 ----------------
ws_d = wb.create_sheet("일별배분")
ws_d.append(["날짜", "요일번호", "요일", "구분", "기준지수(자사몰 요일평균)", "제외조정지수", "배분금액(원)"])
style_header(ws_d, 1, 7)

start = date(2026, 8, 1)
n_days = 31
first_row = 2
last_row = first_row + n_days - 1

for i in range(n_days):
    d = start + timedelta(days=i)
    row = first_row + i
    ws_d.cell(row=row, column=1, value=d).number_format = DATE_FMT
    ws_d.cell(row=row, column=1).font = BLUE
    ws_d.cell(row=row, column=2, value=f"=WEEKDAY(A{row},1)-1").font = BLACK
    ws_d.cell(row=row, column=3, value=f'=CHOOSE(B{row}+1,"일","월","화","수","목","금","토")').font = BLACK

    is_fixed_formula = f"=IF(A{row}=가정!$B$4,\"지그재그 라이브 고정\",\"요일비중 배분\")"
    ws_d.cell(row=row, column=4, value=is_fixed_formula).font = BLACK

    idx_formula = f"=INDEX('요일별비중'!$C$2:$C$8,MATCH(B{row},'요일별비중'!$A$2:$A$8,0))"
    c5 = ws_d.cell(row=row, column=5, value=idx_formula)
    c5.font = BLACK
    c5.number_format = WON_FMT

    excl_idx_formula = f"=IF(A{row}=가정!$B$4,0,E{row})"
    c6 = ws_d.cell(row=row, column=6, value=excl_idx_formula)
    c6.font = BLACK
    c6.number_format = WON_FMT

    if i < n_days - 1:
        amt_formula = (
            f"=IF(A{row}=가정!$B$4,가정!$B$5,"
            f"ROUND(F{row}/SUM($F${first_row}:$F${last_row})*(가정!$B$2-가정!$B$5),0))"
        )
    else:
        # last day absorbs the rounding remainder among non-fixed days
        amt_formula = f"=가정!$B$2-SUM(G{first_row}:G{last_row - 1})"
    c7 = ws_d.cell(row=row, column=7, value=amt_formula)
    c7.font = BLACK
    c7.number_format = WON_FMT

sum_row2 = last_row + 1
ws_d.cell(row=sum_row2, column=3, value="합계").font = BOLD
tot2 = ws_d.cell(row=sum_row2, column=7, value=f"=SUM(G{first_row}:G{last_row})")
tot2.font = BOLD
tot2.number_format = WON_FMT

check_row = sum_row2 + 1
ws_d.cell(row=check_row, column=3, value="목표 대비 차액(검증, 0이어야 함)").font = BLACK
diff_cell = ws_d.cell(row=check_row, column=7, value=f"=G{sum_row2}-가정!$B$2")
diff_cell.font = BLACK
diff_cell.number_format = WON_FMT

autofit(ws_d, [14, 10, 8, 20, 22, 16, 18])
ws_d.freeze_panes = "A2"

wb._sheets = [ws_a, ws_d, ws_w, ws_r, ws_c]
wb.calculation.fullCalcOnLoad = True
wb.save(OUT_PATH)
print("saved", OUT_PATH)
